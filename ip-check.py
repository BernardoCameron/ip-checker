#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# IP Threat Intelligence Enricher
# Fuentes: VirusTotal, AbuseIPDB, Shodan, IPInfo
# Entrada:  .txt (una IP por linea) o .xlsx/.csv
# Salida:   Excel con colores + JSON + reporte HTML

import os
import sys
import io
import json
import time
import argparse
import requests
from datetime import datetime
from pathlib import Path

# Fix Windows console UTF-8
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")


# ─────────────────────────────────────────────
#  Carga de .env
# ─────────────────────────────────────────────
def load_dotenv():
    """Lee un archivo .env en el directorio actual y carga las variables."""
    env_path = Path(".env")
    if not env_path.exists():
        return
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


load_dotenv()

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


# ─────────────────────────────────────────────
#  CONFIGURACIÓN DE APIs
# ─────────────────────────────────────────────
API_KEYS = {
    "virustotal": os.getenv("VT_API_KEY", ""),
    "abuseipdb": os.getenv("ABUSEIPDB_API_KEY", ""),
    "shodan": os.getenv("SHODAN_API_KEY", ""),
    # IPInfo tiene tier gratuito sin key (50k req/mes)
    "ipinfo": os.getenv("IPINFO_TOKEN", ""),
}

RATE_LIMIT_DELAY = 1.2  # segundos entre requests (free tier, se puede modificar)


# ─────────────────────────────────────────────
#  FUENTES
# ─────────────────────────────────────────────


def query_virustotal(ip: str) -> dict:
    result = {"source": "VirusTotal", "status": "skipped"}
    if not API_KEYS["virustotal"]:
        result["status"] = "no_api_key"
        return result
    try:
        url = f"https://www.virustotal.com/api/v3/ip_addresses/{ip}"
        headers = {"x-apikey": API_KEYS["virustotal"]}
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code == 200:
            data = r.json().get("data", {}).get("attributes", {})
            stats = data.get("last_analysis_stats", {})
            result = {
                "source": "VirusTotal",
                "status": "ok",
                "malicious": stats.get("malicious", 0),
                "suspicious": stats.get("suspicious", 0),
                "harmless": stats.get("harmless", 0),
                "undetected": stats.get("undetected", 0),
                "reputation": data.get("reputation", 0),
                "network": data.get("network", ""),
                "country": data.get("country", ""),
                "as_owner": data.get("as_owner", ""),
                "categories": list(data.get("categories", {}).values())[:3],
                "vt_link": f"https://www.virustotal.com/gui/ip-address/{ip}",
            }
        elif r.status_code == 401:
            result["status"] = "invalid_api_key"
        elif r.status_code == 429:
            result["status"] = "rate_limited"
        else:
            result["status"] = f"http_{r.status_code}"
    except Exception as e:
        result["status"] = f"error: {e}"
    return result


def query_abuseipdb(ip: str) -> dict:
    result = {"source": "AbuseIPDB", "status": "skipped"}
    if not API_KEYS["abuseipdb"]:
        result["status"] = "no_api_key"
        return result
    try:
        url = "https://api.abuseipdb.com/api/v2/check"
        headers = {"Key": API_KEYS["abuseipdb"], "Accept": "application/json"}
        params = {"ipAddress": ip, "maxAgeInDays": 90, "verbose": True}
        r = requests.get(url, headers=headers, params=params, timeout=10)
        if r.status_code == 200:
            data = r.json().get("data", {})
            result = {
                "source": "AbuseIPDB",
                "status": "ok",
                "abuse_score": data.get("abuseConfidenceScore", 0),
                "total_reports": data.get("totalReports", 0),
                "last_reported": data.get("lastReportedAt", ""),
                "country_code": data.get("countryCode", ""),
                "isp": data.get("isp", ""),
                "domain": data.get("domain", ""),
                "is_tor": data.get("isTor", False),
                "is_public": data.get("isPublic", True),
                "usage_type": data.get("usageType", ""),
                "abuse_link": f"https://www.abuseipdb.com/check/{ip}",
            }
        elif r.status_code == 401:
            result["status"] = "invalid_api_key"
        elif r.status_code == 429:
            result["status"] = "rate_limited"
        else:
            result["status"] = f"http_{r.status_code}"
    except Exception as e:
        result["status"] = f"error: {e}"
    return result


def query_shodan(ip: str) -> dict:
    result = {"source": "Shodan", "status": "skipped"}
    if not API_KEYS["shodan"]:
        result["status"] = "no_api_key"
        return result
    try:
        url = f"https://api.shodan.io/shodan/host/{ip}"
        params = {"key": API_KEYS["shodan"]}
        r = requests.get(url, params=params, timeout=10)
        if r.status_code == 200:
            data = r.json()
            ports = data.get("ports", [])
            vulns = list(data.get("vulns", {}).keys())
            hostnames = data.get("hostnames", [])
            result = {
                "source": "Shodan",
                "status": "ok",
                "ports": ports[:15],
                "open_ports": len(ports),
                "hostnames": hostnames[:3],
                "org": data.get("org", ""),
                "os": data.get("os", ""),
                "isp": data.get("isp", ""),
                "vulnerabilities": vulns[:5],
                "vuln_count": len(vulns),
                "last_update": data.get("last_update", ""),
                "tags": data.get("tags", []),
                "shodan_link": f"https://www.shodan.io/host/{ip}",
            }
        elif r.status_code == 404:
            result["status"] = "not_found"
        elif r.status_code == 401:
            result["status"] = "invalid_api_key"
        elif r.status_code == 429:
            result["status"] = "rate_limited"
        else:
            result["status"] = f"http_{r.status_code}"
    except Exception as e:
        result["status"] = f"error: {e}"
    return result


def query_ipinfo(ip: str) -> dict:
    result = {"source": "IPInfo", "status": "skipped"}
    try:
        token = API_KEYS["ipinfo"]
        url = f"https://ipinfo.io/{ip}/json"
        params = {"token": token} if token else {}
        r = requests.get(url, params=params, timeout=10)
        if r.status_code == 200:
            data = r.json()
            lat, lon = "", ""
            if "loc" in data:
                parts = data["loc"].split(",")
                if len(parts) == 2:
                    lat, lon = parts[0], parts[1]
            result = {
                "source": "IPInfo",
                "status": "ok",
                "city": data.get("city", ""),
                "region": data.get("region", ""),
                "country": data.get("country", ""),
                "org": data.get("org", ""),
                "hostname": data.get("hostname", ""),
                "timezone": data.get("timezone", ""),
                "latitude": lat,
                "longitude": lon,
                "privacy": data.get("privacy", {}),
                "abuse_contact": data.get("abuse", {}).get("email", ""),
            }
        else:
            result["status"] = f"http_{r.status_code}"
    except Exception as e:
        result["status"] = f"error: {e}"
    return result


# ─────────────────────────────────────────────
#  CARGA DE IPs
# ─────────────────────────────────────────────


def load_ips(filepath: str) -> list[str]:
    # Carga IPs desde .txt, .csv o .xlsx.
    path = Path(filepath)
    ips = []

    if not path.exists():
        print(f"[ERROR] Archivo no encontrado: {filepath}")
        sys.exit(1)

    ext = path.suffix.lower()

    if ext == ".txt":
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                ip = line.strip()
                if ip and not ip.startswith("#"):
                    ips.append(ip)

    elif ext in (".xlsx", ".xls") and PANDAS_AVAILABLE:
        df = pd.read_excel(path, header=None)
        for col in df.columns:
            for val in df[col].dropna():
                ip = str(val).strip()
                if ip:
                    ips.append(ip)
        ips = list(dict.fromkeys(ips))  # deduplicar preservando orden

    elif ext == ".csv" and PANDAS_AVAILABLE:
        df = pd.read_csv(path, header=None)
        for col in df.columns:
            for val in df[col].dropna():
                ip = str(val).strip()
                if ip:
                    ips.append(ip)
        ips = list(dict.fromkeys(ips))

    else:
        print(f"[ERROR] Formato no soportado o pandas no instalado: {ext}")
        sys.exit(1)

    # Filtro basico de formato IP
    valid = []
    for ip in ips:
        parts = ip.split(".")
        if len(parts) == 4:
            try:
                if all(0 <= int(p) <= 255 for p in parts):
                    valid.append(ip)
            except ValueError:
                pass
        if ip not in valid:
            print(f"  [WARN] Ignorando entrada no valida: {ip!r}")

    print(f"[+] {len(valid)} IPs validas cargadas desde {path.name}")
    return valid


# ─────────────────────────────────────────────
#  ENRIQUECIMIENTO PRINCIPAL
# ─────────────────────────────────────────────


def enrich_ip(ip: str, sources: list[str]) -> dict:
    # Consulta todas las fuentes
    print(f"  [~] Enriqueciendo {ip}...", end=" ", flush=True)
    result = {
        "ip": ip,
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "sources": {},
    }

    source_map = {
        "virustotal": query_virustotal,
        "abuseipdb": query_abuseipdb,
        "shodan": query_shodan,
        "ipinfo": query_ipinfo,
    }

    tags = []
    for src in sources:
        fn = source_map.get(src)
        if fn:
            data = fn(ip)
            result["sources"][src] = data
            time.sleep(RATE_LIMIT_DELAY)

    # Calcular veredicto consolidado
    score = 0
    vt = result["sources"].get("virustotal", {})
    ab = result["sources"].get("abuseipdb", {})
    sh = result["sources"].get("shodan", {})

    if vt.get("status") == "ok":
        score += vt.get("malicious", 0) * 3
        score += vt.get("suspicious", 0) * 1

    if ab.get("status") == "ok":
        abuse = ab.get("abuse_score", 0)
        if abuse > 80:
            score += 10
        elif abuse > 50:
            score += 5
        elif abuse > 20:
            score += 2
        if ab.get("is_tor"):
            score += 3

    if sh.get("status") == "ok":
        score += min(sh.get("vuln_count", 0), 5) * 2

    if score >= 15:
        verdict = "CRITICO"
    elif score >= 8:
        verdict = "ALTO"
    elif score >= 3:
        verdict = "MEDIO"
    elif score > 0:
        verdict = "BAJO"
    else:
        verdict = "LIMPIO"

    result["verdict"] = verdict
    result["risk_score"] = score

    print(f"[{verdict}]")
    return result


# ─────────────────────────────────────────────
#  EXCEL
# ─────────────────────────────────────────────

VERDICT_COLORS = {
    "CRITICO": "C62828",
    "ALTO": "E65100",
    "MEDIO": "F9A825",
    "BAJO": "558B2F",
    "LIMPIO": "1565C0",
}


def export_excel(results: list[dict], output_path: str):
    if not EXCEL_AVAILABLE:
        print("[WARN] openpyxl no instalado, omitiendo Excel.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IP Enrichment"

    # colores de cabecera segun fuente
    header_colors = {
        "General": "263238",
        "VirusTotal": "1565C0",
        "AbuseIPDB": "4A148C",
        "Shodan": "BF360C",
        "IPInfo": "1B5E20",
    }

    columns = [
        # (header, key path, source group)
        ("IP", "ip", "General"),
        ("Veredicto", "verdict", "General"),
        ("Risk Score", "risk_score", "General"),
        ("Timestamp", "timestamp", "General"),
        # VT
        ("VT Malicious", "sources.virustotal.malicious", "VirusTotal"),
        ("VT Suspicious", "sources.virustotal.suspicious", "VirusTotal"),
        ("VT Reputation", "sources.virustotal.reputation", "VirusTotal"),
        ("VT AS Owner", "sources.virustotal.as_owner", "VirusTotal"),
        ("VT Country", "sources.virustotal.country", "VirusTotal"),
        ("VT Link", "sources.virustotal.vt_link", "VirusTotal"),
        # AbuseIPDB
        ("Abuse Score", "sources.abuseipdb.abuse_score", "AbuseIPDB"),
        ("Abuse Reports", "sources.abuseipdb.total_reports", "AbuseIPDB"),
        ("ISP", "sources.abuseipdb.isp", "AbuseIPDB"),
        ("Usage Type", "sources.abuseipdb.usage_type", "AbuseIPDB"),
        ("Is TOR", "sources.abuseipdb.is_tor", "AbuseIPDB"),
        ("Last Reported", "sources.abuseipdb.last_reported", "AbuseIPDB"),
        ("Abuse Link", "sources.abuseipdb.abuse_link", "AbuseIPDB"),
        # Shodan
        ("Open Ports", "sources.shodan.open_ports", "Shodan"),
        ("Ports", "sources.shodan.ports", "Shodan"),
        ("Vulns", "sources.shodan.vulnerabilities", "Shodan"),
        ("Vuln Count", "sources.shodan.vuln_count", "Shodan"),
        ("OS", "sources.shodan.os", "Shodan"),
        ("Shodan Link", "sources.shodan.shodan_link", "Shodan"),
        # IPInfo
        ("City", "sources.ipinfo.city", "IPInfo"),
        ("Region", "sources.ipinfo.region", "IPInfo"),
        ("Country", "sources.ipinfo.country", "IPInfo"),
        ("Org", "sources.ipinfo.org", "IPInfo"),
        ("Timezone", "sources.ipinfo.timezone", "IPInfo"),
        ("Latitude", "sources.ipinfo.latitude", "IPInfo"),
        ("Longitude", "sources.ipinfo.longitude", "IPInfo"),
    ]

    # Escribir cabeceras con color por grupo
    for col_idx, (header, _, group) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        hex_color = header_colors.get(group, "263238")
        cell.fill = PatternFill("solid", fgColor=hex_color)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    ws.row_dimensions[1].height = 30

    def get_nested(d, path):
        # Navega un dict por path tipo "sources.virustotal.malicious"
        keys = path.split(".")
        for k in keys:
            if isinstance(d, dict):
                d = d.get(k, "")
            else:
                return ""
        if isinstance(d, list):
            return ", ".join(str(x) for x in d)
        return d if d is not None else ""

    # Escribir filas de datos
    for row_idx, result in enumerate(results, 2):
        verdict = result.get("verdict", "LIMPIO")
        row_color = {
            "CRITICO": "FFEBEE",
            "ALTO": "FFF3E0",
            "MEDIO": "FFFDE7",
            "BAJO": "F1F8E9",
            "LIMPIO": "E3F2FD",
        }.get(verdict, "FFFFFF")

        for col_idx, (_, key_path, _) in enumerate(columns, 1):
            value = get_nested(result, key_path)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = PatternFill("solid", fgColor=row_color)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.font = Font(size=10)

            # Celdas veredicto
            if key_path == "verdict":
                vcolor = VERDICT_COLORS.get(verdict, "000000")
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill("solid", fgColor=vcolor)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Links
            if (
                "link" in key_path
                and isinstance(value, str)
                and value.startswith("http")
            ):
                cell.hyperlink = value
                cell.font = Font(color="1565C0", underline="single", size=10)

    # Cols
    col_widths = {
        "IP": 16,
        "Veredicto": 12,
        "Risk Score": 10,
        "Timestamp": 22,
        "VT Malicious": 12,
        "VT Suspicious": 13,
        "VT Reputation": 13,
        "VT AS Owner": 22,
        "VT Country": 10,
        "VT Link": 18,
        "Abuse Score": 12,
        "Abuse Reports": 14,
        "ISP": 25,
        "Usage Type": 18,
        "Is TOR": 8,
        "Last Reported": 22,
        "Abuse Link": 18,
        "Open Ports": 11,
        "Ports": 30,
        "Vulns": 30,
        "Vuln Count": 11,
        "OS": 14,
        "Shodan Link": 18,
        "City": 16,
        "Region": 16,
        "Country": 10,
        "Org": 28,
        "Timezone": 22,
        "Latitude": 10,
        "Longitude": 10,
    }
    for col_idx, (header, _, _) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(
            header, 14
        )

    # Freeze top row
    ws.freeze_panes = "A2"

    # Hoja resumen
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Resumen de analisis"
    ws2["A1"].font = Font(bold=True, size=14)

    counts = {}
    for r in results:
        v = r.get("verdict", "LIMPIO")
        counts[v] = counts.get(v, 0) + 1

    ws2["A3"] = "Total IPs analizadas"
    ws2["B3"] = len(results)
    for i, (verdict, count) in enumerate(sorted(counts.items()), 4):
        ws2[f"A{i}"] = verdict
        ws2[f"B{i}"] = count
        hex_color = VERDICT_COLORS.get(verdict, "000000")
        ws2[f"A{i}"].font = Font(bold=True, color="FFFFFF")
        ws2[f"A{i}"].fill = PatternFill("solid", fgColor=hex_color)
        ws2[f"B{i}"].fill = PatternFill("solid", fgColor=hex_color)
        ws2[f"B{i}"].font = Font(bold=True, color="FFFFFF")

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12

    ws2["A" + str(len(counts) + 6)] = (
        f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    )

    wb.save(output_path)
    print(f"[+] Excel guardado: {output_path}")


# ─────────────────────────────────────────────
# HTML
# ─────────────────────────────────────────────


def export_html(results: list[dict], output_path: str):
    verdict_badge = {
        "CRITICO": ("#C62828", "#FFEBEE"),
        "ALTO": ("#E65100", "#FFF3E0"),
        "MEDIO": ("#F9A825", "#FFFDE7"),
        "BAJO": ("#558B2F", "#F1F8E9"),
        "LIMPIO": ("#1565C0", "#E3F2FD"),
    }

    rows_html = ""
    for r in results:
        v = r.get("verdict", "LIMPIO")
        score = r.get("risk_score", 0)
        col, bg = verdict_badge.get(v, ("#000", "#fff"))
        ip = r["ip"]

        vt = r["sources"].get("virustotal", {})
        ab = r["sources"].get("abuseipdb", {})
        sh = r["sources"].get("shodan", {})
        geo = r["sources"].get("ipinfo", {})

        vt_mal = vt.get("malicious", "N/A") if vt.get("status") == "ok" else "N/A"
        ab_scr = ab.get("abuse_score", "N/A") if ab.get("status") == "ok" else "N/A"
        ab_rep = ab.get("total_reports", "N/A") if ab.get("status") == "ok" else "N/A"
        sh_ports = len(sh.get("ports", [])) if sh.get("status") == "ok" else "N/A"
        sh_vulns = sh.get("vuln_count", "N/A") if sh.get("status") == "ok" else "N/A"
        city = (
            f"{geo.get('city','')}, {geo.get('country','')}"
            if geo.get("status") == "ok"
            else "N/A"
        )
        isp = ab.get("isp", geo.get("org", "N/A")) or "N/A"
        is_tor = "SI" if ab.get("is_tor") else "No"

        vt_link = vt.get("vt_link", "#")
        ab_link = ab.get("abuse_link", "#")
        sh_link = sh.get("shodan_link", "#")

        rows_html += f"""
        <tr style="background:{bg}">
          <td><code style="font-size:13px">{ip}</code></td>
          <td><span class="badge" style="background:{col};color:#fff">{v}</span></td>
          <td style="text-align:center;font-weight:600">{score}</td>
          <td style="text-align:center;color:{'#C62828' if isinstance(vt_mal,int) and vt_mal>0 else '#555'}">{vt_mal}</td>
          <td style="text-align:center;color:{'#C62828' if isinstance(ab_scr,int) and ab_scr>50 else '#555'}">{ab_scr}</td>
          <td style="text-align:center">{ab_rep}</td>
          <td style="text-align:center">{sh_ports}</td>
          <td style="text-align:center;color:{'#C62828' if isinstance(sh_vulns,int) and sh_vulns>0 else '#555'}">{sh_vulns}</td>
          <td>{city}</td>
          <td style="font-size:12px;max-width:180px;overflow:hidden;text-overflow:ellipsis">{isp}</td>
          <td style="text-align:center;color:{'#C62828' if is_tor=='SI' else '#555'}">{is_tor}</td>
          <td>
            {'<a href="'+vt_link+'" target="_blank">VT</a> ' if vt_link!='#' else ''}
            {'<a href="'+ab_link+'" target="_blank">Abuse</a> ' if ab_link!='#' else ''}
            {'<a href="'+sh_link+'" target="_blank">Shodan</a>' if sh_link!='#' else ''}
          </td>
        </tr>"""

    summary = {}
    for r in results:
        v = r.get("verdict", "LIMPIO")
        summary[v] = summary.get(v, 0) + 1

    summary_html = ""
    for v, c in sorted(summary.items()):
        col, bg = verdict_badge.get(v, ("#000", "#eee"))
        summary_html += f'<div class="stat-card" style="border-top:4px solid {col}"><div class="stat-num">{c}</div><div class="stat-label">{v}</div></div>'

    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>IP Threat Intelligence Report</title>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ font-family: -apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif; background:#f5f5f5; color:#212121; }}
  .header {{ background:#1a237e; color:#fff; padding:28px 40px; }}
  .header h1 {{ font-size:22px; font-weight:600; }}
  .header p  {{ font-size:13px; opacity:.75; margin-top:4px; }}
  .content {{ padding:28px 40px; max-width:1400px; margin:0 auto; }}
  .summary {{ display:flex; gap:14px; margin-bottom:28px; flex-wrap:wrap; }}
  .stat-card {{ background:#fff; border-radius:8px; padding:16px 22px; min-width:110px; box-shadow:0 1px 3px rgba(0,0,0,.1); }}
  .stat-num {{ font-size:28px; font-weight:700; }}
  .stat-label {{ font-size:12px; color:#666; margin-top:2px; }}
  .table-wrap {{ overflow-x:auto; background:#fff; border-radius:8px; box-shadow:0 1px 3px rgba(0,0,0,.1); }}
  table {{ width:100%; border-collapse:collapse; font-size:13px; }}
  thead th {{ background:#263238; color:#fff; padding:10px 12px; text-align:left; font-weight:500; white-space:nowrap; }}
  tbody tr {{ border-bottom:1px solid rgba(0,0,0,.06); transition:filter .15s; }}
  tbody tr:hover {{ filter:brightness(.96); }}
  td {{ padding:9px 12px; vertical-align:middle; }}
  .badge {{ font-size:11px; font-weight:600; padding:3px 8px; border-radius:4px; white-space:nowrap; }}
  a {{ color:#1565C0; text-decoration:none; margin-right:4px; }}
  a:hover {{ text-decoration:underline; }}
  .footer {{ text-align:center; padding:20px; font-size:12px; color:#999; }}
</style>
</head>
<body>
<div class="header">
  <h1>IP Threat Intelligence Report</h1>
  <p>Generado: {now} &nbsp;|&nbsp; Total IPs: {len(results)} &nbsp;|&nbsp; Fuentes: VirusTotal, AbuseIPDB, Shodan, IPInfo</p>
</div>
<div class="content">
  <div class="summary">{summary_html}</div>
  <div class="table-wrap">
    <table>
      <thead>
        <tr>
          <th>IP</th><th>Veredicto</th><th>Score</th>
          <th>VT Malicious</th><th>Abuse Score</th><th>Abuse Reports</th>
          <th>Puertos</th><th>CVEs</th>
          <th>Ubicacion</th><th>ISP / Org</th><th>TOR</th><th>Links</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>
</div>
<div class="footer">IP Threat Intelligence Checker &mdash; github.com/BernardoCameron/ip-checker</div>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8", newline="\n") as f:
        f.write(html)
    print(f"[+] HTML guardado: {output_path}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="IP Threat Intelligence Checker",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python ip-check.py ips.txt
  python ip-check.py ips.xlsx --sources virustotal abuseipdb
  python ip-check.py ips.txt --sources all --output reporte

Variables de entorno (API keys):
  VT_API_KEY       VirusTotal
  ABUSEIPDB_API_KEY
  SHODAN_API_KEY
  IPINFO_TOKEN     (opcional, free tier sin key)
        """,
    )
    parser.add_argument("input", help="Archivo de IPs (.txt, .xlsx, .csv)")
    parser.add_argument(
        "--sources",
        nargs="+",
        choices=["virustotal", "abuseipdb", "shodan", "ipinfo", "all"],
        default=["all"],
        help="Fuentes a consultar (default: all)",
    )
    parser.add_argument(
        "--output", default="ip_report", help="Nombre base del output (sin extensión)"
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=1.2,
        help="Segundos entre requests (default: 1.2)",
    )
    args = parser.parse_args()

    global RATE_LIMIT_DELAY
    RATE_LIMIT_DELAY = args.delay

    sources = (
        ["virustotal", "abuseipdb", "shodan", "ipinfo"]
        if "all" in args.sources
        else args.sources
    )

    print("\n+======================================+")
    print("|  IP Threat Intelligence Checker     |")
    print("+======================================+\n")

    # Verificar keys configuradas
    print("[*] API keys configuradas:")
    for src in sources:
        key = API_KEYS.get(src, "")
        status = "OK" if key else "NO CONFIGURADA (se omitira)"
        if src == "ipinfo":
            status = f"{'token configurado' if key else 'free tier (sin token)'}"
        print(f"    {src:<15} {status}")
    print()

    ips = load_ips(args.input)
    if not ips:
        print("[ERROR] No se encontraron IPs validas.")
        sys.exit(1)

    print(
        f"\n[*] Iniciando enriquecimiento de {len(ips)} IPs con: {', '.join(sources)}\n"
    )

    results = []
    for i, ip in enumerate(ips, 1):
        print(f"[{i}/{len(ips)}] ", end="")
        result = enrich_ip(ip, sources)
        results.append(result)

    # Exportar resultados
    print(f"\n[*] Exportando resultados...")

    json_path = f"{args.output}.json"
    with open(json_path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"[+] JSON guardado: {json_path}")

    if EXCEL_AVAILABLE:
        export_excel(results, f"{args.output}.xlsx")

    export_html(results, f"{args.output}.html")

    # Resumen final en consola
    print("\n" + "-" * 45)
    print("RESUMEN:")
    counts = {}
    for r in results:
        v = r.get("verdict", "LIMPIO")
        counts[v] = counts.get(v, 0) + 1
    for verdict in ["CRITICO", "ALTO", "MEDIO", "BAJO", "LIMPIO"]:
        if verdict in counts:
            print(f"  {verdict:<10} {counts[verdict]} IP(s)")
    print("-" * 45 + "\n")


if __name__ == "__main__":
    main()
